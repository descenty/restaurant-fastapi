import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from core.config import settings

from .schemas import DishCascadeCreate, MenuCascadeCreate, SubmenuCascadeCreate

# id_map: dict[int, UUID] = {}


def parse_menus():
    wb = openpyxl.load_workbook(settings.menus_xl_path)
    sheet: Worksheet = wb.active  # type: ignore
    menus: list[MenuCascadeCreate] = []
    try:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            if row[0] is not None:
                menus.append(MenuCascadeCreate.model_validate(row))
            elif row[1] is not None:
                menus[-1].submenus.append(SubmenuCascadeCreate.model_validate(row[1:]))
            elif row[2] is not None:
                menus[-1].submenus[-1].dishes.append(
                    DishCascadeCreate.model_validate(row[2:])
                )
    except ValidationError as e:
        print(e.errors()[0]['msg'])
    return menus
