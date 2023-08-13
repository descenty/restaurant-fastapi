from decimal import Decimal

# from uuid import UUID
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError, model_validator

from core.config import settings
from schemas.dish import DishCreate
from schemas.menu import MenuCreate
from schemas.submenu import SubmenuCreate


class DishCascadeCreate(DishCreate):
    id: int

    @model_validator(mode='before')
    def validator(cls, row: tuple):
        id, title, description, price, discount = row[0], row[1], row[2], row[3], row[4]
        assert isinstance(id, int) and id > 0, 'id must be positive integer'
        assert len(title) > 0, 'title must be non-empty string'
        assert len(description) > 0, 'description must be non-empty string'
        assert (
            isinstance(price, float) and (d_price := Decimal(price)) > 0
        ), 'price must be positive decimal'
        assert (
            isinstance(discount, float) and (d_discount := Decimal(discount)) >= 0
        ), 'discount must be non-negative decimal'
        return {
            'id': id,
            'title': title,
            'description': description,
            'price': d_price,
            'discount': d_discount,
        }


class SubmenuCascadeCreate(SubmenuCreate):
    id: int
    dishes: list[DishCreate] = []

    @model_validator(mode='before')
    def validator(cls, row: tuple):
        return MenuCascadeCreate.model_validate(row).model_dump()


class MenuCascadeCreate(MenuCreate):
    id: int
    submenus: list[SubmenuCascadeCreate] = []

    @model_validator(mode='before')
    def validator(cls, v: tuple):
        id, title, description = v[0], v[1], v[2]
        assert isinstance(id, int) and id > 0, 'id must be positive integer'
        assert len(title) > 0, 'title must be non-empty string'
        assert len(description) > 0, 'description must be non-empty string'
        return {'id': id, 'title': title, 'description': description}


# id_map: dict[int, UUID] = {}


def parse_menus():
    wb = openpyxl.load_workbook(settings.menus_xl_path)
    sheet: Worksheet = wb.active  # type: ignore
    menus: list[MenuCascadeCreate] = []
    try:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            if row[0] is not None:
                menus.append(MenuCascadeCreate.model_validate(row))
            elif row[1] is not None:
                menus[-1].submenus.append(SubmenuCascadeCreate.model_validate(row[1:]))
            elif row[2] is not None:
                menus[-1].submenus[-1].dishes.append(
                    DishCascadeCreate.model_validate(row[2:])
                )
    except ValidationError as e:
        print(e.errors()[0]['msg'])
    return menus
